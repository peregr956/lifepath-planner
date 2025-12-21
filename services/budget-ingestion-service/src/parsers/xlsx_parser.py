from __future__ import annotations

from io import BytesIO
from typing import Dict, Iterable, List, Optional, Sequence

from models.raw_budget import DraftBudgetModel, RawBudgetLine
from parsers.format_detection import HeaderSignals, detect_format

CategoryHeaders = ("category", "category_label", "category name", "type")
AmountHeaders = ("amount", "value", "total")
DescriptionHeaders = ("description", "memo", "notes", "note")
DateHeaders = ("date", "transaction date", "posted date")
LedgerDebitHeaders = ("debit", "withdrawal", "withdraw", "dr")
LedgerCreditHeaders = ("credit", "deposit", "cr")
LedgerBalanceHeaders = ("balance", "running balance")


def parse_xlsx_to_draft_model(file_bytes: bytes) -> DraftBudgetModel:
    """Parse an XLSX upload into a DraftBudgetModel skeleton."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - defensive guard
        raise RuntimeError("openpyxl is required for XLSX parsing.") from exc

    workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return DraftBudgetModel(
            detected_format="unknown",
            notes="XLSX file is empty.",
        )

    headers = _normalize_headers(rows[0])
    category_key = _find_column(headers, CategoryHeaders)
    amount_key = _find_column(headers, AmountHeaders)
    description_key = _find_column(headers, DescriptionHeaders)
    date_key = _find_column(headers, DateHeaders)
    header_signals = _extract_header_signals(headers)

    warnings: List[str] = []
    if category_key is None:
        warnings.append("Category column not detected; leaving labels empty.")
    if amount_key is None:
        warnings.append("Amount column not detected; skipping lines without numeric value.")

    lines: List[RawBudgetLine] = []
    for row_index, raw_row in enumerate(rows[1:], start=2):
        row = {header: value for header, value in zip(headers, raw_row)}

        category_value = row.get(category_key, "").strip() if category_key else ""
        amount_value = row.get(amount_key) if amount_key else None

        if amount_value is None or str(amount_value).strip() == "":
            continue

        amount = _parse_amount(amount_value)
        description = row.get(description_key) if description_key else None
        description = description.strip() if isinstance(description, str) else description
        parsed_date = _parse_date(row.get(date_key)) if date_key else None
        metadata = _build_metadata(row, kept_keys={category_key, amount_key, description_key, date_key})

        lines.append(
            RawBudgetLine(
                source_row_index=row_index,
                date=parsed_date,
                category_label=category_value,
                description=description,
                amount=amount,
                metadata=metadata,
            )
        )

    detected_format, format_hints = detect_format(lines, header_signals)
    combined_notes = " ".join(warnings) if warnings else None
    return DraftBudgetModel(
        lines=lines,
        detected_format=detected_format,
        notes=combined_notes,
        format_hints=format_hints,
    )


def _normalize_headers(header_row: Sequence[object]) -> List[str]:
    normalized: List[str] = []
    for cell in header_row:
        if cell is None:
            normalized.append("")
        else:
            normalized.append(str(cell).strip())
    return normalized


def _find_column(headers: Sequence[str], candidates: Iterable[str]) -> Optional[str]:
    lowered = {header.lower().strip(): header for header in headers if header}
    for candidate in candidates:
        normalized = candidate.lower().strip()
        if normalized in lowered:
            return lowered[normalized]
    return None


def _parse_amount(raw_value: object) -> float:
    if raw_value is None:
        return 0.0
    if isinstance(raw_value, (int, float)):
        return float(raw_value)

    cleaned = str(raw_value).replace(",", "").strip()
    for symbol in ("$", "€", "£"):
        cleaned = cleaned.replace(symbol, "")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _parse_date(raw_value: object):
    from datetime import datetime

    if not raw_value:
        return None

    if hasattr(raw_value, "date"):
        return raw_value.date()

    text = str(raw_value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _build_metadata(row: Dict[str, object], kept_keys: Iterable[Optional[str]]) -> Dict[str, object]:
    reserved = {key for key in kept_keys if key}
    return {key: value for key, value in row.items() if key not in reserved and value not in (None, "")}


def _extract_header_signals(headers: Sequence[str]) -> HeaderSignals:
    normalized = {header.strip().lower() for header in headers if header}
    return HeaderSignals(
        has_debit_column=any(candidate in normalized for candidate in LedgerDebitHeaders),
        has_credit_column=any(candidate in normalized for candidate in LedgerCreditHeaders),
        has_balance_column=any(candidate in normalized for candidate in LedgerBalanceHeaders),
    )
